"""Read a spreadsheet as a WORKBOOK, not as a rectangle of values.

`pandas.read_excel` and every tool built on it opens an .xlsx, evaluates
nothing, keeps the cached values, and throws away the formulas, the hidden
rows, the merged ranges and the calculation state before a human or a model
ever sees the file. Six of the defects below are invisible after that call, by
construction, no matter how good the model reading the dataframe is:

    a total whose range stops above the last row of its own column
    a formula column with three constants pasted over it
    a #REF! that a chart is quietly plotting as zero
    eleven hidden rows that a SUM includes and a reader does not
    merged cells that shifted every row below them by one
    formulas that have never been calculated at all

The most expensive spreadsheet mistake in economics was the first of those.
Reinhart and Rogoff's 2010 growth-and-debt result averaged L30:L44 in a column
whose data ran to row 49, silently excluding five countries; the paper was
cited in budget debates on two continents before Herndon, Ash and Pollin found
it in 2013. No amount of reading the numbers finds that. Reading the formula
finds it immediately.

This module needs openpyxl, which the core does not. It is an optional extra,
exactly like Pillow for S15, and every check here reports UNAVAILABLE with the
install line rather than passing quietly when it is absent. A check that cannot
see must never look like a check that saw nothing wrong.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}

# Excel's error values. A cell holding one of these is not "missing data": it is
# a calculation that failed and was then saved, charted, and reported.
ERROR_VALUES = {
    "#REF!": "a reference to a deleted cell or sheet",
    "#DIV/0!": "a division by zero",
    "#VALUE!": "an operand of the wrong type",
    "#NAME?": "an unrecognised function or defined name",
    "#NULL!": "an empty intersection of two ranges",
    "#NUM!": "a numeric operation with no valid result",
    "#N/A": "a lookup that found nothing",
    "#SPILL!": "a dynamic array with nowhere to spill",
    "#CALC!": "a calculation engine error",
    "#GETTING_DATA": "an external query that never returned",
}

# Aggregates whose range is a claim about coverage. A SUM that stops short is a
# wrong total; an IF that stops short is usually just an IF.
COVERAGE_FUNCTIONS = {"SUM", "AVERAGE", "AVERAGEA", "COUNT", "COUNTA", "MIN", "MAX",
                      "MEDIAN", "PRODUCT", "STDEV", "STDEVP", "VAR", "SUBTOTAL",
                      "SUMPRODUCT", "LARGE", "SMALL"}

RANGE_RE = re.compile(r"(?<![A-Za-z0-9_!])\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")
FUNCTION_RE = re.compile(r"([A-Z][A-Z0-9.]*)\s*\(", re.IGNORECASE)

MAX_EXAMPLES = 6
MAX_SHEET_CELLS = 2_000_000     # a guard on pathological files, not a feature
MIN_FORMULA_COLUMN = 3          # below this, "a formula column" is not a pattern


class WorkbookUnavailable(RuntimeError):
    """openpyxl is not installed, and every X check must say so out loud."""


UNAVAILABLE_REASON = (
    "reading a workbook's structure needs openpyxl, which is not installed, so "
    "formulas, hidden rows, merged cells and calculation state cannot be looked "
    "at. Install it with: pip install 'dinostomp[xlsx]'"
)


def available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def looks_like_workbook(path: str | Path) -> bool:
    return Path(path).suffix.lower() in WORKBOOK_SUFFIXES


def _require():
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised by the skip path
        raise WorkbookUnavailable(UNAVAILABLE_REASON) from exc
    return openpyxl


class Sheet:
    """One worksheet, read twice: once for formulas, once for cached values.

    Both passes are needed and neither is sufficient. The formula pass is the
    only place a range, a constant pasted over a formula, or an uncalculated
    cell exists. The value pass is the only place an error value that a chart
    is plotting exists. A tool that opens the file once has already chosen
    which half of the defects it cannot find.
    """

    def __init__(self, name: str, formulas, values, dims, merged, state: str):
        self.name = name
        self.formulas = formulas          # {(row, col): "=SUM(A1:A9)" | constant}
        self.values = values              # {(row, col): cached value}
        self.hidden_rows, self.hidden_cols = dims
        self.merged = merged
        self.state = state                # visible | hidden | veryHidden

    @property
    def max_row(self) -> int:
        return max((r for r, _ in self.formulas), default=0)

    def populated_rows_in_column(self, col: int) -> set[int]:
        return {r for (r, c), v in self.formulas.items()
                if c == col and v is not None and str(v).strip() != ""}


def load_sheets(path: str | Path) -> list[Sheet]:
    """Load every worksheet, formulas and cached values together."""
    openpyxl = _require()
    path = Path(path)
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets: list[Sheet] = []
    try:
        for name in wb_f.sheetnames:
            ws_f, ws_v = wb_f[name], wb_v[name]
            if ws_f.max_row * max(1, ws_f.max_column) > MAX_SHEET_CELLS:
                continue
            formulas, values = {}, {}
            for row in ws_f.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        formulas[(cell.row, cell.column)] = cell.value
            for row in ws_v.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        values[(cell.row, cell.column)] = cell.value
            hidden_rows = {i for i, d in ws_f.row_dimensions.items() if d.hidden}
            hidden_cols = {k for k, d in ws_f.column_dimensions.items() if d.hidden}
            merged = [str(r) for r in ws_f.merged_cells.ranges]
            sheets.append(Sheet(name, formulas, values, (hidden_rows, hidden_cols),
                                merged, ws_f.sheet_state))
    finally:
        wb_f.close()
        wb_v.close()
    return sheets


def sheet_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[dict], list[str]]:
    """The first worksheet as rows of {header: value}, for the G checks.

    Notes are returned rather than printed, because a caller that hides which
    sheet it read is a caller reporting on a file the user did not open.
    """
    openpyxl = _require()
    wb = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    notes: list[str] = []
    try:
        name = sheet_name or wb.sheetnames[0]
        if len(wb.sheetnames) > 1:
            notes.append(f"workbook has {len(wb.sheetnames)} sheets "
                         f"({', '.join(wb.sheetnames[:5])}); read {name!r} for the value checks, "
                         f"structure checks read all of them")
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return [], notes + ["sheet is empty"]
        columns = [str(h).strip() if h is not None else f"column_{i + 1}"
                   for i, h in enumerate(header)]
        blank_headers = sum(1 for h in header if h is None)
        if blank_headers:
            notes.append(f"{blank_headers} column(s) have no header and were named by position")
        rows = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            rows.append({columns[i]: v for i, v in enumerate(values) if i < len(columns)})
        return rows, notes
    finally:
        wb.close()


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _col_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _col_index(letters: str) -> int:
    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - 64)
    return index


Result = tuple[bool, str, int, list[str], dict]


def check_formula_errors(sheets: list[Sheet]) -> Result:
    """XL2: a saved calculation that failed. Gates.

    Deterministic and never benign: `#REF!` means a reference the file itself
    cannot resolve. Charts plot these as zero and SUM ranges propagate them, so
    a single one can move a total without appearing in it.

    Read from the CACHED values, which is where an error actually lives: a
    formula is only `#DIV/0!` once something evaluated it. A workbook written
    by a library and never opened has no cached values at all, so this check
    finds nothing in one, correctly. XL6 is the check that notices that state.
    """
    hits: dict[str, Counter] = defaultdict(Counter)
    examples: list[str] = []
    total = 0
    cells = 0
    for sheet in sheets:
        cells += len(sheet.values)
        for (row, col), value in sheet.values.items():
            text = value if isinstance(value, str) else ""
            if text in ERROR_VALUES:
                hits[sheet.name][text] += 1
                total += 1
                if len(examples) < MAX_EXAMPLES:
                    formula = sheet.formulas.get((row, col))
                    shown = f" from {formula}" if _is_formula(formula) else ""
                    examples.append(f"{sheet.name}!{_col_letter(col)}{row}: {text} "
                                    f"({ERROR_VALUES[text]}){shown}")
    if not total:
        return True, "no error values in any sheet", cells, [], {}
    kinds = Counter()
    for counter in hits.values():
        kinds.update(counter)
    return (False, f"{total} error cell(s) saved in the workbook: "
                   f"{', '.join(f'{k} x{v}' for k, v in kinds.most_common())}",
            cells, examples, {"by_sheet": {k: dict(v) for k, v in hits.items()}})


def check_short_ranges(sheets: list[Sheet]) -> Result:
    """XL5: an aggregate whose range stops above the last row of its own column.

    The Reinhart-Rogoff defect. Gates, and the reasoning matches R9: an
    aggregate is a summary, and a summary that does not cover its own records
    is mechanically wrong rather than debatable.

    The false positive to avoid is the legitimate subtotal, so a row is only
    orphaned if NO aggregate over that column covers it. A block of subtotals
    that between them cover every row reports nothing.
    """
    covered: dict[tuple[str, int], set[int]] = defaultdict(set)
    aggregates: dict[tuple[str, int], list[tuple[str, str]]] = defaultdict(list)
    n_formulas = 0
    for sheet in sheets:
        for (row, col), value in sheet.formulas.items():
            if not _is_formula(value):
                continue
            n_formulas += 1
            functions = {m.group(1).upper() for m in FUNCTION_RE.finditer(value)}
            if not (functions & COVERAGE_FUNCTIONS):
                continue
            for match in RANGE_RE.finditer(value):
                c1, r1, c2, r2 = match.group(1), int(match.group(2)), match.group(3), int(match.group(4))
                if _col_index(c1) != _col_index(c2):
                    continue  # a rectangular range is not a column claim
                target = (sheet.name, _col_index(c1))
                covered[target].update(range(min(r1, r2), max(r1, r2) + 1))
                aggregates[target].append((f"{_col_letter(col)}{row}", value))
    if not aggregates:
        return (True, "no column aggregates to check", n_formulas, [],
                {"not_applicable": "this workbook has no column aggregate to check; "
                                   "a range that stops short needs a range"})

    orphans: dict[tuple[str, int], set[int]] = {}
    for target, rows_covered in covered.items():
        sheet = next(s for s in sheets if s.name == target[0])
        populated = sheet.populated_rows_in_column(target[1])
        # Only rows BELOW the range are the defect: a header above it is normal,
        # and so is a label row. The aggregate's own cell never counts.
        floor = min(rows_covered)
        missed = {r for r in populated if r > max(rows_covered)}
        missed -= {int(ref[1:]) for ref, _ in aggregates[target]
                   if ref[1:].isdigit()}
        missed = {r for r in missed
                  if sheet.formulas.get((r, target[1])) is not None
                  and not _is_formula(sheet.formulas[(r, target[1])])}
        del floor
        if missed:
            orphans[target] = missed
    if not orphans:
        return (True, f"{len(aggregates)} column aggregate(s), each covering every populated "
                      f"row of its column", n_formulas, [], {})
    examples = []
    for (sheet_name, col), missed in list(orphans.items())[:MAX_EXAMPLES]:
        ref, formula = aggregates[(sheet_name, col)][0]
        shown = ", ".join(str(r) for r in sorted(missed)[:5])
        examples.append(f"{sheet_name}!{ref} = {formula} excludes populated "
                        f"{_col_letter(col)} row(s) {shown}"
                        f"{' and more' if len(missed) > 5 else ''}")
    total_missed = sum(len(v) for v in orphans.values())
    return (False, f"{len(orphans)} aggregate range(s) stop above the last populated row of "
                   f"their own column, excluding {total_missed} row(s) from a total that "
                   f"presents itself as complete",
            n_formulas, examples,
            {"orphans": {f"{s}!{_col_letter(c)}": sorted(v) for (s, c), v in orphans.items()}})


def check_pasted_constants(sheets: list[Sheet]) -> Result:
    """XL1: constants sitting in a column that is otherwise formulas.

    Someone pasted a value over a calculation. The cell keeps showing the right
    number and stops tracking its inputs forever, which is why this survives
    every review that reads the rendered sheet.
    """
    hits: dict[str, list[str]] = defaultdict(list)
    n_cells = 0
    _any_formula_column = False
    for sheet in sheets:
        by_column: dict[int, list[tuple[int, Any]]] = defaultdict(list)
        for (row, col), value in sheet.formulas.items():
            by_column[col].append((row, value))
            n_cells += 1
        for col, cells in by_column.items():
            formulas = [(r, v) for r, v in cells if _is_formula(v)]
            if len(formulas) < MIN_FORMULA_COLUMN:
                continue
            _any_formula_column = True
            rows_with_formula = {r for r, _ in formulas}
            span = range(min(rows_with_formula), max(rows_with_formula) + 1)
            constants = [(r, v) for r, v in cells
                         if r in span and not _is_formula(v) and str(v).strip() != ""]
            for row, value in constants[:4]:
                hits[sheet.name].append(
                    f"{_col_letter(col)}{row} = {value!r} sits inside a formula column "
                    f"({len(formulas)} formulas, rows {min(rows_with_formula)}"
                    f"-{max(rows_with_formula)})")
    total = sum(len(v) for v in hits.values())
    if not total:
        return (True, "no constants pasted inside a formula column", n_cells, [],
                {} if _any_formula_column else
                {"not_applicable": "no column in this workbook is mostly formulas, so "
                                   "there is no formula column to paste over"})
    examples = [f"{sheet}: {item}" for sheet, items in hits.items() for item in items]
    return (False, f"{total} constant(s) pasted into formula column(s) across "
                   f"{len(hits)} sheet(s); each stopped tracking its inputs when it was pasted",
            n_cells, examples[:MAX_EXAMPLES], {"by_sheet": {k: len(v) for k, v in hits.items()}})


def check_hidden_cells(sheets: list[Sheet]) -> Result:
    """XL3: rows, columns and sheets a reader will not see and a total will.

    Not a defect on its own, which is why it warns: hiding a working column is
    ordinary. It is reported because the combination of "hidden" and "included
    in an aggregate" is invisible from every direction at once.
    """
    hidden_rows = {s.name: sorted(s.hidden_rows) for s in sheets if s.hidden_rows}
    hidden_cols = {s.name: sorted(s.hidden_cols) for s in sheets if s.hidden_cols}
    hidden_sheets = [s.name for s in sheets if s.state != "visible"]
    if not hidden_rows and not hidden_cols and not hidden_sheets:
        return True, "nothing hidden in this workbook", len(sheets), [], {}
    examples = [f"{name}: {len(rows)} hidden row(s) ({', '.join(str(r) for r in rows[:8])})"
                for name, rows in hidden_rows.items()]
    examples += [f"{name}: hidden column(s) {', '.join(cols[:8])}"
                 for name, cols in hidden_cols.items()]
    examples += [f"sheet {name!r} is {dict((s.name, s.state) for s in sheets)[name]}"
                 for name in hidden_sheets]
    parts = []
    if hidden_rows:
        parts.append(f"{sum(len(v) for v in hidden_rows.values())} hidden row(s)")
    if hidden_cols:
        parts.append(f"{sum(len(v) for v in hidden_cols.values())} hidden column(s)")
    if hidden_sheets:
        parts.append(f"{len(hidden_sheets)} hidden sheet(s)")
    return (False, "; ".join(parts) + ". Aggregates still count what a reader cannot see",
            len(sheets), examples[:MAX_EXAMPLES],
            {"rows": hidden_rows, "columns": hidden_cols, "sheets": hidden_sheets})


def check_merged_cells(sheets: list[Sheet]) -> Result:
    """XL4: merged ranges, which are a layout decision the data pays for.

    A merged range holds its value in the top-left cell and None everywhere
    else. Every reader that turns the sheet into rows gets one populated cell
    and N-1 blanks, which is how a merged header silently shifts a column.
    """
    hits = {s.name: s.merged for s in sheets if s.merged}
    if not hits:
        return True, "no merged cells", len(sheets), [], {}
    total = sum(len(v) for v in hits.values())
    examples = [f"{name}: {', '.join(ranges[:6])}" for name, ranges in hits.items()]
    return (False, f"{total} merged range(s) across {len(hits)} sheet(s); each reads as one "
                   f"value and a row of blanks to anything that imports this file",
            len(sheets), examples[:MAX_EXAMPLES], {"by_sheet": hits})


def check_uncalculated(sheets: list[Sheet]) -> Result:
    """XL6: formulas the file has never calculated.

    openpyxl reads the value Excel last cached. When that is absent the cell has
    a formula and no result, so anything reading values sees a blank where a
    number belongs. This does NOT claim to detect a STALE value: proving a
    cached number disagrees with its formula needs a calculation engine, and
    saying so is more useful than pretending otherwise.
    """
    hits: dict[str, int] = defaultdict(int)
    examples: list[str] = []
    n_formulas = 0
    for sheet in sheets:
        for (row, col), value in sheet.formulas.items():
            if not _is_formula(value):
                continue
            n_formulas += 1
            if sheet.values.get((row, col)) is None:
                hits[sheet.name] += 1
                if len(examples) < MAX_EXAMPLES:
                    examples.append(f"{sheet.name}!{_col_letter(col)}{row} = {value} "
                                    f"has no cached result")
    if not n_formulas:
        return (True, "no formulas in this workbook", 0, [],
                {"not_applicable": "this workbook contains no formulas to calculate"})
    total = sum(hits.values())
    if not total:
        return (True, f"all {n_formulas} formula(s) carry a cached result", n_formulas, [], {})
    if total == n_formulas:
        # EVERY formula uncached means no spreadsheet application has ever
        # opened this file: it was written by a library. That is a fact about
        # the file's provenance, not a defect in someone's model, and reporting
        # it as one would fire on every programmatically generated workbook and
        # teach people to ignore this check. The partial case below is the one
        # that means something.
        return (True, f"none of the {n_formulas} formula(s) carry a cached result, so this "
                      f"workbook has never been opened by a spreadsheet application. Nothing "
                      f"is stale, and nothing is calculated: a reader of VALUES sees blanks "
                      f"where every formula is",
                n_formulas, examples[:2], {"never_opened": True})
    return (False, f"{total} of {n_formulas} formula(s) have never been calculated while the "
                   f"rest have; anything reading values sees a blank where a number belongs "
                   f"(a cached value that is merely STALE cannot be detected without a "
                   f"calculation engine, and is not claimed here)",
            n_formulas, examples, {"by_sheet": dict(hits)})
