"""The table arm of DinoTrials: planted defects in spreadsheets, not evals.

The main harness builds POD trials and its own header records that wiring a
dataset arm into the scorecard "is the right fix and is not done". The G and XL
series need exactly that arm, because a check whose only evidence is its own
unit test has been graded by the person who wrote it, once.

Same rubric as the pod arm, both directions:

    SENSITIVITY  a file with exactly one thing wrong, and the id and level the
                 battery must produce for it
    SPECIFICITY  files with nothing wrong, which must produce zero findings

The clean controls are deliberately awkward rather than sterile. `orders.csv`
carries a legitimately repeating order id (a line-item table, the most common
spreadsheet shape there is) and real decimals over 999. `subtotals.xlsx`
carries a block of SUBTOTAL formulas that between them cover every row, which
is the shape XL5 must never mistake for a short range. A battery that stays
quiet only on sterile data has not been tested for specificity at all.
"""

from __future__ import annotations

import csv
from pathlib import Path

CLEAN_HEADER = ["order_id", "line_no", "vendor", "region", "qty", "unit_price", "ship_date"]
CLEAN_ROWS = [
    ["ORD-1001", "1", "Acme Corp", "West", "10", "1200.00", "2026-01-05"],
    ["ORD-1001", "2", "Acme Corp", "West", "4", "300.50", "2026-01-05"],
    ["ORD-1001", "3", "Acme Corp", "West", "7", "88.25", "2026-01-05"],
    ["ORD-1002", "1", "Globex", "East", "5", "640.00", "2026-01-06"],
    ["ORD-1003", "1", "Initech", "North", "12", "980.10", "2026-01-07"],
    ["ORD-1004", "1", "Umbrella Ltd", "South", "3", "150.75", "2026-01-08"],
    ["ORD-1005", "1", "Soylent", "West", "8", "220.00", "2026-01-09"],
    ["ORD-1006", "1", "Hooli", "East", "6", "410.40", "2026-01-10"],
]
XL_HEADER = ["order_id", "vendor", "qty", "unit_price"]
XL_ROWS = [
    ["ORD-1001", "Acme Corp", 10, 1200.00],
    ["ORD-1002", "Globex", 4, 300.50],
    ["ORD-1003", "Initech", 7, 88.25],
    ["ORD-1004", "Umbrella Ltd", 5, 640.00],
    ["ORD-1005", "Soylent", 12, 980.10],
    ["ORD-1006", "Hooli", 3, 150.75],
]


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _csv(root: Path, rows, header=None, name="table.csv") -> Path:
    path = root / name
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(header or CLEAN_HEADER)
        writer.writerows(rows)
    return path


def _rows():
    return [row[:] for row in CLEAN_ROWS]


def _xlsx(root: Path, *, formulas=None, hidden_rows=(), hidden_cols=(), merged=(),
          hidden_sheet=False, name="book.xlsx") -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(XL_HEADER)
    for row in XL_ROWS:
        ws.append(list(row))
    for ref, value in (formulas or {}).items():
        ws[ref] = value
    for index in hidden_rows:
        ws.row_dimensions[index].hidden = True
    for letter in hidden_cols:
        ws.column_dimensions[letter].hidden = True
    for ref in merged:
        ws.merge_cells(ref)
    if hidden_sheet:
        wb.create_sheet("Scratch").sheet_state = "hidden"
    path = root / name
    wb.save(path)
    return path


# ---- sensitivity: one planted defect each ------------------------------


def t_invisible_character(root: Path) -> Path:
    rows = _rows()
    rows[2][2] = "Acme Corp"      # a non-breaking space, invisible everywhere
    return _csv(root, rows)


def t_duplicate_row(root: Path) -> Path:
    return _csv(root, _rows() + [CLEAN_ROWS[3][:]])


def t_violated_identifier(root: Path) -> Path:
    header = ["invoice_id", "vendor", "amount"]
    rows = [[f"INV-{i:04d}", "Acme", str(100 + i)] for i in range(12)]
    rows[7][0] = rows[2][0]
    return _csv(root, rows, header)


def t_leading_zeros(root: Path) -> Path:
    header = ["zip", "city"]
    rows = [["00501", "Holtsville"], ["02134", "Boston"], ["07001", "Avenel"],
            ["08540", "Princeton"], ["10001", "New York"], ["01887", "Wilmington"]]
    return _csv(root, rows, header)


def t_mixed_date_formats(root: Path) -> Path:
    rows = _rows()
    rows[4][6] = "07/01/2026"
    return _csv(root, rows)


def t_text_in_a_numeric_column(root: Path) -> Path:
    rows = _rows()
    rows[5][4] = "n/a"
    return _csv(root, rows)


def t_repeated_numeric_sentinel(root: Path) -> Path:
    rows = _rows()
    for i in (1, 3, 5, 7):
        rows[i][5] = "999999"
    return _csv(root, rows)


def t_category_split_by_case(root: Path) -> Path:
    rows = _rows()
    rows[3][3] = "east"
    rows[6][3] = "WEST"
    return _csv(root, rows)


def t_two_percent_scales(root: Path) -> Path:
    header = ["vendor", "margin_pct"]
    rows = [["Acme", "0.15"], ["Globex", "0.20"], ["Initech", "18"],
            ["Umbrella", "0.11"], ["Soylent", "22"], ["Hooli", "0.09"]]
    return _csv(root, rows, header)


def t_currency_formatted_numbers(root: Path) -> Path:
    rows = _rows()
    for row in rows:
        row[5] = f"${float(row[5]):,.2f}"
    return _csv(root, rows)


def t_near_duplicate_row(root: Path) -> Path:
    near = CLEAN_ROWS[3][:]
    near[2] = near[2].upper()
    return _csv(root, _rows() + [near])


def t_pasted_constant(root: Path) -> Path:
    formulas = {f"E{i}": f"=C{i}*D{i}" for i in range(2, 8)}
    formulas["E5"] = 4321
    return _xlsx(root, formulas=formulas)


def t_saved_error_value(root: Path) -> Path:
    return _xlsx(root, formulas={"D8": "#REF!"})


def t_hidden_rows(root: Path) -> Path:
    return _xlsx(root, hidden_rows=(3,), hidden_cols=("D",), hidden_sheet=True)


def t_merged_cells(root: Path) -> Path:
    return _xlsx(root, merged=("A9:C9",))


def t_total_stops_short(root: Path) -> Path:
    """The Reinhart-Rogoff defect: a SUM whose range ends above its own data."""
    return _xlsx(root, formulas={"C9": "=SUM(C2:C5)"})


# ---- specificity: nothing wrong, nothing reported ----------------------


def clean_line_item_csv(root: Path) -> Path:
    """A repeating order id is what a line-item table IS, not a duplicate key."""
    return _csv(root, _rows())


def clean_subtotal_workbook(root: Path) -> Path:
    """Two subtotals that between them cover every populated row. The shape XL5
    must never mistake for a range that stops short."""
    return _xlsx(root, formulas={"C9": "=SUM(C2:C4)", "C10": "=SUM(C5:C7)"})


TABLE_TRIALS = [
    ("a non-breaking space inside a vendor name", t_invisible_character, ("G1", "warn")),
    ("a byte-identical duplicate row", t_duplicate_row, ("G2", "fail")),
    ("an identifier column with one repeat", t_violated_identifier, ("G3", "warn")),
    ("zip codes whose leading zeros a conversion would eat", t_leading_zeros, ("G4", "warn")),
    ("one date column written two ways", t_mixed_date_formats, ("G5", "warn")),
    ("'n/a' inside a quantity column", t_text_in_a_numeric_column, ("G6", "warn")),
    ("999999 repeated in a price column", t_repeated_numeric_sentinel, ("G7", "warn")),
    ("a region column split by capitalisation", t_category_split_by_case, ("G8", "warn")),
    ("a rate column holding 0.15 and 18", t_two_percent_scales, ("G9", "warn")),
    ("prices stored as $1,200.00", t_currency_formatted_numbers, ("G10", "warn")),
    ("a row identical but for capitalisation", t_near_duplicate_row, ("G11", "warn")),
    ("a constant pasted over a formula", t_pasted_constant, ("XL1", "warn"), openpyxl_available),
    ("a #REF! saved in the workbook", t_saved_error_value, ("XL2", "fail"), openpyxl_available),
    ("rows, a column and a sheet hidden from the reader", t_hidden_rows, ("XL3", "warn"),
     openpyxl_available),
    ("a merged range that flattens a row on import", t_merged_cells, ("XL4", "warn"),
     openpyxl_available),
    ("a total whose range stops above its own data", t_total_stops_short, ("XL5", "fail"),
     openpyxl_available),
]

CLEAN_TABLE_TRIALS = [
    ("a line-item table with a repeating order id", clean_line_item_csv, None),
    ("a workbook whose subtotals cover every row", clean_subtotal_workbook, openpyxl_available),
]


# ---- the join arm: two tables, one planted defect between them ---------

JOIN_CHILD_HEADER = ["order_id", "vendor_code", "amount"]
JOIN_CHILD_ROWS = [
    ["ORD-1001", "ACME", "120.00"],
    ["ORD-1002", "GLOBEX", "300.50"],
    ["ORD-1003", "INITECH", "88.25"],
    ["ORD-1004", "ACME", "640.00"],
    ["ORD-1005", "HOOLI", "980.10"],
    ["ORD-1006", "GLOBEX", "150.75"],
]
JOIN_PARENT_HEADER = ["vendor_code", "vendor_name", "amount"]
JOIN_PARENT_ROWS = [
    ["ACME", "Acme Corp", "760.00"],
    ["GLOBEX", "Globex", "451.25"],
    ["INITECH", "Initech", "88.25"],
    ["HOOLI", "Hooli", "980.10"],
    ["SOYLENT", "Soylent", "0.00"],
]


def _pair(root: Path, child=None, parent=None) -> tuple[Path, Path]:
    left = _csv(root, child or JOIN_CHILD_ROWS, JOIN_CHILD_HEADER, "orders.csv")
    right = _csv(root, parent or JOIN_PARENT_ROWS, JOIN_PARENT_HEADER, "vendors.csv")
    return left, right


def j_nothing_matches(root: Path):
    parent = [["X-" + r[0], r[1], r[2]] for r in JOIN_PARENT_ROWS]
    return _pair(root, parent=parent)


def j_orphan_rows(root: Path):
    return _pair(root, child=JOIN_CHILD_ROWS + [["ORD-1007", "UMBRELLA", "12.00"]])


def j_case_only_mismatch(root: Path):
    """The ACNH defect: one capital letter, no error anywhere."""
    child = [r[:] for r in JOIN_CHILD_ROWS]
    child[2][1] = "initech"
    return _pair(root, child=child)


def j_parent_key_repeats(root: Path):
    return _pair(root, parent=JOIN_PARENT_ROWS + [["ACME", "Acme Corp (old)", "0.00"]])


def j_key_type_mismatch(root: Path):
    child = [[f"ORD-{i}", f"0{i}0", "10.00"] for i in range(1, 7)]
    parent = [[str(int(f"0{i}0")), f"Vendor {i}", "10.00"] for i in range(1, 7)]
    return _pair(root, child=child, parent=parent)


def j_total_disagrees(root: Path):
    parent = [r[:] for r in JOIN_PARENT_ROWS]
    parent[1][2] = "999.99"
    return _pair(root, parent=parent)


def clean_join(root: Path):
    """A line-item table against a vendor lookup: the ordinary shape."""
    return _pair(root)


# A trial may name the keys. Two of these defects are only reachable that way,
# and that is a property of the defects rather than a convenience: when the
# intended key matches NOTHING, or when the two sides store it as different
# types, there is nothing for inference to find. Saying which columns you meant
# is the only way to be told they are broken.
JOIN_TRIALS = [
    ("an inner join that returns nothing at all", j_nothing_matches, ("JN1", "fail"),
     ("vendor_code", "vendor_code")),
    ("a code the lookup table has never heard of", j_orphan_rows, ("JN2", "warn")),
    ("one capital letter between the two tables", j_case_only_mismatch, ("JN3", "fail")),
    ("a lookup key that appears twice", j_parent_key_repeats, ("JN4", "warn")),
    ("a join that multiplies its own rows", j_parent_key_repeats, ("JN5", "warn")),
    ("the same key stored as text and as a number", j_key_type_mismatch, ("JN6", "warn"),
     ("vendor_code", "vendor_code")),
    ("a parent total that disagrees with its own children", j_total_disagrees, ("JN7", "fail")),
]

CLEAN_JOIN_TRIALS = [
    ("a line-item table against its vendor lookup", clean_join, None),
]
